from datetime import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings

from admin_models.models import Users
from .models import Oders, Payment_Status, Payment_Service, Baskets
# Create your views here.
from openpyxl import Workbook, load_workbook
import pathlib


def add_order(order_id, price, payment_service, address_deliver):
    # Get current date and time
    base_dir = settings.BASE_DIR
    exel_file = f'exel/orders{datetime.now().strftime("%d-%m-%Y")}.xlsx'
    path_to_exel_file = pathlib.Path(base_dir) / exel_file

    # Create or load the Excel file
    try:
        wb = load_workbook(filename=path_to_exel_file)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        # Set the headers in the Excel file
        ws.append(["Время оплаты", "Номер заказа", "Цена", "Сервис оплаты", "Адрес доставки"])

    # Add the order data to the Excel file
    ws.append([datetime.now().strftime("%H:%M:%S"), order_id, price, payment_service, address_deliver,])

    # Save the changes and close the Excel file
    wb.save(filename=path_to_exel_file)
    wb.close()

def order_redirect(request, order_id):
    try:
        order: Oders = Oders.objects.get(order_id=order_id)
    except ObjectDoesNotExist:
        return HttpResponse("Данного заказа не сущестует")

    # Здесь должна быть подготовка и отправка запросов на платежную систему

    return HttpResponse(f"Заказ:{order_id}, который переходит к {order.payment.name} способу оплаты"
                        f'<a href="/orders/{order_id}/complete">Для перехода к "оплате" нажмите сюда</a>')


def order_compete(request, order_id):
    try:
        order: Oders = Oders.objects.get(order_id=order_id)
    except ObjectDoesNotExist:
        return HttpResponse("Данного заказа не сущестует")

    payment_status_paid = Payment_Status.objects.get(name="paid")
    payment_status_wait = Payment_Status.objects.get(name="wait")


    # Здесь должна быть проверка действительно ли пользователь оплатил заказ
    if order.payment_status == payment_status_wait:
        order.payment_status = payment_status_paid
        order.save()
        add_order(order_id, order.price, order.payment.name, order.address_deliver)
        basket: Baskets = order.basket
        basket.sell = True
        user = Users.objects.get(id=basket.user.id)
        basket.save()
        new_basket = Baskets.objects.create(user=user)

        return HttpResponse(f"Спасибо что оплатили ваш заказ!")

    return HttpResponse(f"Заказ уже обработан")
